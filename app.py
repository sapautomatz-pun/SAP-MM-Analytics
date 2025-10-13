# ==========================================================
# SAP AUTOMATZ – Executive Procurement Analytics
# Version: v33.1 (Fixes: Streamlit deprecation, PDF watermark/text, Risk breakdown)
# ==========================================================

import os, io, re, datetime, math
import pandas as pd, numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import streamlit as st
from openai import OpenAI
from fpdf import FPDF
from unidecode import unidecode

# ------------------------- CONFIG -------------------------
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
MODEL = "gpt-4o-mini"
LOGO_URL = "https://raw.githubusercontent.com/sapautomatz-pun/SAP-MM-Analytics/1d3346d7d35396f13ff06da26f24ebb5ebb70f23/sapautomatz_logo.png"

client = OpenAI(api_key=OPENAI_API_KEY)

# ------------------------- STREAMLIT PAGE -------------------------
st.set_page_config(page_title="SAP Automatz – Executive Procurement Analytics", page_icon="📊", layout="wide")
st.markdown("<style>.stApp header{visibility:hidden}</style>", unsafe_allow_html=True)

col1, col2 = st.columns([1,3])
with col1:
    st.image(LOGO_URL, width=140)
with col2:
    st.markdown("""
        <h2 style='margin-bottom:0;color:#1a237e;font-size:26px;'>SAP Automatz – AI Procurement Analytics</h2>
        <p style='color:#444;margin-top:0;font-size:14px;'>ERP-Compatible Executive Dashboard<br>
        <b>Automate. Analyze. Accelerate.</b></p>
    """, unsafe_allow_html=True)
st.divider()

# ------------------------- HELPERS -------------------------
def sanitize_text(t):
    if t is None: return ""
    return unidecode(str(t))

def parse_amount_and_currency(v, fallback="INR"):
    if pd.isna(v): return 0.0, fallback
    if isinstance(v,(int,float,np.number)): return float(v), fallback
    s=str(v)
    sym_map={"₹":"INR","Rs":"INR","$":"USD","USD":"USD","€":"EUR","EUR":"EUR"}
    cur=fallback
    for sym,c in sym_map.items():
        if sym in s:
            cur=c
            s=s.replace(sym,"")
    s=re.sub(r"[^\d.\-]", "", s)
    try:
        amt=float(s)
    except:
        amt=0.0
    return amt, cur

def clean_dataframe(df):
    # Ensure currency col
    if "CURRENCY" not in df.columns:
        df["CURRENCY"] = "INR"
    amounts=[]; currencies=[]
    for _, row in df.iterrows():
        amt, cur = parse_amount_and_currency(row.get("AMOUNT", 0), row.get("CURRENCY", "INR"))
        amounts.append(amt)
        currencies.append(cur)
    df["AMOUNT_NUM"] = amounts
    df["CURRENCY_DETECTED"] = currencies
    if "VENDOR" in df.columns:
        df["VENDOR"] = df["VENDOR"].astype(str).fillna("Unknown")
    if "MATERIAL" in df.columns:
        df["MATERIAL"] = df["MATERIAL"].astype(str).fillna("Unknown")
    return df

def compute_kpis(df):
    df = clean_dataframe(df)
    if "PO_DATE" in df.columns:
        df["PO_DATE"] = pd.to_datetime(df["PO_DATE"], errors="coerce")
    totals = df.groupby("CURRENCY_DETECTED")["AMOUNT_NUM"].sum().to_dict()
    total_spend = float(sum(totals.values())) if totals else 0.0
    dominant = max(totals, key=totals.get) if totals else None
    top_v = df.groupby("VENDOR")["AMOUNT_NUM"].sum().nlargest(10).to_dict() if "VENDOR" in df.columns else {}
    if "QUANTITY" in df.columns:
        top_m = df.groupby("MATERIAL")["QUANTITY"].sum().nlargest(10).to_dict()
    else:
        top_m = df.groupby("MATERIAL")["AMOUNT_NUM"].sum().nlargest(10).to_dict() if "MATERIAL" in df.columns else {}
    monthly = {}
    if "PO_DATE" in df.columns:
        d = df.dropna(subset=["PO_DATE"])
        if not d.empty:
            d["YM"] = d["PO_DATE"].dt.to_period("M").astype(str)
            monthly = d.groupby("YM")["AMOUNT_NUM"].sum().to_dict()
    return {"totals": totals, "total_spend": total_spend, "dominant": dominant,
            "top_v": top_v, "top_m": top_m, "monthly": monthly, "records": len(df), "df": df}

# Procurement risk same as earlier version (kept explainable)
def compute_procurement_risk(df, kpis):
    df_local = kpis.get("df", df)
    totals = kpis.get("totals", {})
    total_spend = kpis.get("total_spend", 0.0)
    vendor_sums = df_local.groupby("VENDOR")["AMOUNT_NUM"].sum()
    num_vendors = vendor_sums.size if not vendor_sums.empty else 0
    top_vendor_share = (vendor_sums.max() / total_spend) if total_spend and not vendor_sums.empty else 1.0
    vendor_concentration_score = max(0.0, (1.0 - top_vendor_share)) * 100
    vendor_diversity_score = min(100.0, (num_vendors / 50.0) * 100.0)
    if totals and total_spend:
        dominant = kpis.get("dominant")
        dominant_share = totals.get(dominant, 0.0) / total_spend if dominant else 1.0
        currency_exposure_score = dominant_share * 100.0
    else:
        currency_exposure_score = 100.0
    monthly = list(kpis.get("monthly", {}).values())
    if len(monthly) >= 3 and np.mean(monthly) > 0:
        cv = np.std(monthly) / (np.mean(monthly) + 1e-9)
        monthly_volatility_score = max(0.0, 1.0 - min(cv, 2.0)) * 100.0
    else:
        monthly_volatility_score = 80.0
    w_conc = 0.4; w_div = 0.2; w_curr=0.2; w_vol=0.2
    final_score = (vendor_concentration_score * w_conc +
                   vendor_diversity_score * w_div +
                   currency_exposure_score * w_curr +
                   monthly_volatility_score * w_vol)
    final_score = float(max(0.0, min(100.0, final_score)))
    if final_score >= 67:
        band = "Low"
    elif final_score >= 34:
        band = "Medium"
    else:
        band = "High"
    breakdown = {"vendor_concentration_score": vendor_concentration_score,
                 "vendor_diversity_score": vendor_diversity_score,
                 "currency_exposure_score": currency_exposure_score,
                 "monthly_volatility_score": monthly_volatility_score}
    return {"score": final_score, "band": band, "breakdown": breakdown, "num_vendors": num_vendors, "top_vendor_share": top_vendor_share}

# AI prompt function: increased token budget, sanitized
def generate_ai(kpis, summary_only=False):
    totals_text = "\n".join([f"- {c}: {v:,.2f}" for c, v in kpis["totals"].items()]) if kpis["totals"] else "No currency totals."
    top_v_text = "\n".join([f"{i+1}. {v}: {amt:,.2f}" for i,(v,amt) in enumerate(kpis["top_v"].items())]) if kpis["top_v"] else "No vendor data."
    prompt = f"""
You are a procurement analytics expert. Using the KPIs below, produce:
1) Executive Insights — 4 to 6 concise bullets.
2) Recommendations — 4 concise bullets.
3) Key Action Points — 4 to 6 prioritized actions.

KPI summary:
Total records: {kpis['records']}
Total spend: {kpis['total_spend']:,.2f}
Totals by currency:
{totals_text}
Top vendors:
{top_v_text}
Monthly sample: {', '.join(list(kpis['monthly'].keys())[:6]) if kpis['monthly'] else 'N/A'}

Return the three sections clearly separated and use professional, actionable phrasing.
"""
    try:
        r = client.chat.completions.create(
            model=MODEL,
            messages=[{"role":"system","content":"You are an expert procurement analyst."},{"role":"user","content":prompt}],
            temperature=0.2,
            max_tokens=900
        )
        return sanitize_text(r.choices[0].message.content)
    except Exception as e:
        return f"AI Error: {e}"

# ------------------------- PDF (Unicode-safe) with safer text writing -------------------------
class PDF(FPDF):
    def header(self): pass
    def footer(self):
        self.set_y(-15)
        # ensure font exists; we'll register before using PDF
        try:
            self.set_font("DejaVu", "I", 8)
        except:
            self.set_font("Helvetica", "I", 8)
        self.set_text_color(130,130,130)
        self.cell(0,10,"© 2025 SAP Automatz – Executive Procurement Analytics",align="C")

    # draw watermark very light and drawn BEFORE content usage
    def watermark(self, text="SAP Automatz – Automate • Analyze • Accelerate"):
        # very light gray so it never competes with content
        self.set_text_color(245,245,245)
        try:
            self.set_font("DejaVu", "B", 28)
        except:
            self.set_font("Helvetica", "B", 28)
        # Draw rotation transform, then text, then reset
        self.rotate(45)
        # center-ish placement
        self.text(25, 150, text)
        self.rotate(0)

    def rotate(self, angle):
        if angle != 0:
            self._out(f"q {math.cos(angle*math.pi/180):.6f} {math.sin(angle*math.pi/180):.6f} "
                      f"{-math.sin(angle*math.pi/180):.6f} {math.cos(angle*math.pi/180):.6f} 0 0 cm")
        else:
            self._out("Q")

def add_tile(pdf, x, y, w, h, title, value, color=(21,101,192)):
    pdf.set_fill_color(*color)
    pdf.rect(x,y,w,h,"F")
    pdf.set_text_color(255,255,255)
    pdf.set_xy(x+3, y+3)
    try:
        pdf.set_font("DejaVu", "B", 11)
    except:
        pdf.set_font("Helvetica","B",11)
    pdf.cell(w-6,7, title, ln=True)
    pdf.set_xy(x+3, y+11)
    try:
        pdf.set_font("DejaVu","",10)
    except:
        pdf.set_font("Helvetica","",10)
    pdf.cell(w-6,6, str(value), ln=True)

# safe multi-cell: chunk long lines and fallback to sanitized substring if needed
def safe_multi_cell(pdf, w, h, text):
    # split text into lines by newline, then chunk each long line
    max_chunk = 500
    for paragraph in str(text).split("\n"):
        paragraph = paragraph.strip()
        if not paragraph:
            pdf.ln(2); continue
        # chunk
        i = 0
        while i < len(paragraph):
            chunk = paragraph[i:i+max_chunk]
            try:
                pdf.multi_cell(w, h, chunk)
            except Exception:
                # fallback: try smaller chunk
                try:
                    pdf.multi_cell(w, h, chunk[:200])
                except Exception:
                    # last resort: write sanitized short text
                    pdf.multi_cell(w, h, sanitize_text(chunk)[:150])
                # continue
            i += max_chunk

def generate_pdf(ai_text, kpis, charts, company, summary_text, risk):
    pdf = PDF()

    # Register fonts BEFORE any page operations
    font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    if not os.path.exists(font_path):
        os.makedirs("fonts", exist_ok=True)
        import urllib.request
        urllib.request.urlretrieve(
            "https://github.com/dejavu-fonts/dejavu-fonts/raw/version_2_37/ttf/DejaVuSans.ttf",
            "fonts/DejaVuSans.ttf"
        )
        font_path = "fonts/DejaVuSans.ttf"

    try:
        pdf.add_font("DejaVu", "", font_path, uni=True)
        pdf.add_font("DejaVu", "B", font_path, uni=True)
        pdf.add_font("DejaVu", "I", font_path, uni=True)
    except Exception:
        # ignore if registration fails; FPDF will fallback
        pass

    # -------- COVER PAGE (watermark drawn first so it is behind cover text)
    pdf.add_page()
    pdf.watermark()
    if os.path.exists(LOGO_URL) and LOGO_URL.startswith("http"):
        # FPDF can't directly place remote images; ignore here - we place via local file if available
        pass
    # try to place logo via requests -> local temp if possible
    try:
        r = requests.get(LOGO_URL, timeout=10)
        if r.status_code == 200:
            with open("tmp_logo.png","wb") as f:
                f.write(r.content)
            pdf.image("tmp_logo.png", x=70, y=18, w=70)
    except Exception:
        # ignore logo fetch errors
        pass

    try:
        pdf.set_font("DejaVu", "B", 20)
    except:
        pdf.set_font("Helvetica", "B", 20)
    pdf.ln(75)
    pdf.cell(0,10, "Executive Procurement Analysis Report", align="C", ln=True)
    try:
        pdf.set_font("DejaVu", "", 12)
    except:
        pdf.set_font("Helvetica", "", 12)
    pdf.cell(0,8, f"Prepared for: {company}", align="C", ln=True)
    pdf.cell(0,8, f"Generated on: {datetime.date.today().strftime('%d %B %Y')}", align="C", ln=True)
    pdf.ln(10)
    # summary_text may be long; use safe writing
    try:
        pdf.set_font("DejaVu","",11)
    except:
        pdf.set_font("Helvetica","",11)
    safe_multi_cell(pdf, 0, 7, summary_text)

    # -------- KPI DASHBOARD
    pdf.add_page()
    pdf.watermark()
    pdf.set_fill_color(26,35,126)
    pdf.rect(0,0,210,25,"F")
    pdf.set_text_color(255,255,255)
    try:
        pdf.set_font("DejaVu","B",16)
    except:
        pdf.set_font("Helvetica","B",16)
    pdf.cell(0,15,"Executive Procurement Dashboard", align="C", ln=True)
    pdf.set_text_color(0,0,0)
    pdf.ln(10)

    y = pdf.get_y()
    add_tile(pdf, 10, y, 50, 22, "Total Records", kpis["records"], (13,71,161))
    add_tile(pdf, 65, y, 70, 22, "Total Spend", f"{kpis['total_spend']:,.2f}", (21,101,192))
    add_tile(pdf, 140, y, 60, 22, "Dominant Currency", kpis["dominant"], (30,136,229))
    pdf.ln(36)
    # Risk tile
    color_risk = (56,142,60) if risk["band"]=="Low" else ((242,153,74) if risk["band"]=="Medium" else (192,39,0))
    add_tile(pdf, 10, pdf.get_y(), 85, 22, "Procurement Risk Index", f"{risk['score']:.0f} ({risk['band']})", color_risk)
    pdf.ln(28)

    # -------- AI INSIGHTS
    pdf.add_page()
    pdf.watermark()
    try:
        pdf.set_font("DejaVu","B",13)
    except:
        pdf.set_font("Helvetica","B",13)
    pdf.cell(0,10,"AI-Generated Executive Insights", ln=True)
    try:
        pdf.set_font("DejaVu","",11)
    except:
        pdf.set_font("Helvetica","",11)
    safe_multi_cell(pdf, 0, 7, ai_text)

    # -------- RISK BREAKDOWN PAGE - write as full table lines
    pdf.add_page()
    pdf.watermark()
    try:
        pdf.set_font("DejaVu","B",13)
    except:
        pdf.set_font("Helvetica","B",13)
    pdf.cell(0,10,"Procurement Risk – Breakdown", ln=True)
    pdf.set_font("DejaVu","",11)
    bd = risk["breakdown"]
    for k,v in bd.items():
        pdf.multi_cell(0,7, f"{k.replace('_',' ').title()}: {v:,.2f}")

    # -------- CHARTS PAGES (one chart per page) - ensure enough margin
    for ch in charts:
        if os.path.exists(ch):
            pdf.add_page()
            pdf.watermark()
            try:
                pdf.set_font("DejaVu","B",12)
            except:
                pdf.set_font("Helvetica","B",12)
            title = os.path.basename(ch).replace("_"," ").replace(".png","").title()
            pdf.cell(0,10, title, ln=True)
            # place image slightly down to keep header visible
            try:
                pdf.image(ch, x=18, y=35, w=174)
            except Exception:
                # fallback: small image
                try:
                    pdf.image(ch, x=30, y=40, w=150)
                except:
                    pass

    # final output: use dest="S" and produce bytes safely (latin-1 ignore)
    pdf_bytes = pdf.output(dest="S").encode("latin-1", "ignore")
    return io.BytesIO(pdf_bytes)

# ------------------------- GAUGE CHART (unchanged)
def plot_risk_gauge(score, out_path="gauge_risk.png"):
    fig, ax = plt.subplots(figsize=(6,3))
    ax.axis('off')
    # colored bands
    angles = np.linspace(-np.pi, 0, 100)
    cmap = [(1.0,0.2,0.2), (1.0,0.7,0.2), (0.2,0.7,0.2)]
    splits = [0,33,66,100]
    for i in range(len(splits)-1):
        start = -np.pi + (splits[i]/100.0)*np.pi
        end = -np.pi + (splits[i+1]/100.0)*np.pi
        theta = np.linspace(start, end, 50)
        xs = np.cos(theta); ys = np.sin(theta)
        ax.fill_between(xs, ys, -1.2, color=cmap[i], alpha=0.9)
    theta_score = -np.pi + (score/100.0)*np.pi
    x = 0.9 * math.cos(theta_score); y = 0.9 * math.sin(theta_score)
    ax.plot([0,x],[0,y], lw=4, color='k')
    ax.scatter([0],[0], color='k', s=30)
    ax.text(0, -0.1, f"{score:.0f}", horizontalalignment='center', verticalalignment='center', fontsize=20, fontweight='bold')
    ax.set_xlim(-1.2,1.2); ax.set_ylim(-1.2,0.4)
    plt.axis('off'); fig.savefig(out_path, bbox_inches='tight', dpi=150); plt.close(fig)
    return out_path

# ------------------------- UI FLOW -------------------------
st.title("📊 Executive Procurement Dashboard")
company_name = st.text_input("Enter your Company / Client Name (for report cover):", "ABC Manufacturing Pvt Ltd")

uploaded = st.file_uploader("Upload CSV or XLSX", type=["csv","xlsx"])
if not uploaded:
    st.info("Upload a dataset to start (use the sample datasets).")
    st.stop()

# safe read
try:
    if uploaded.name.lower().endswith(".xlsx"):
        import openpyxl
        df = pd.read_excel(uploaded, engine="openpyxl")
    else:
        df = pd.read_csv(uploaded)
except Exception as e:
    st.error(f"Failed to read file: {e}")
    st.stop()

kpis = compute_kpis(df)
risk = compute_procurement_risk(df, kpis)
gauge_path = plot_risk_gauge(risk['score'], out_path="gauge_risk.png")

# Build charts
charts=[]
if kpis["totals"]:
    fig, ax = plt.subplots(figsize=(5,4))
    labels = list(kpis["totals"].keys()); vals = list(kpis["totals"].values())
    ax.pie(vals, labels=labels, autopct="%1.1f%%", startangle=90); ax.set_title("Currency Distribution")
    fig.tight_layout(); cur_path="chart_currency.png"; fig.savefig(cur_path, bbox_inches='tight', dpi=150); plt.close(fig); charts.append(cur_path)
if kpis["top_v"]:
    fig, ax = plt.subplots(figsize=(8,4))
    vendors = list(kpis["top_v"].keys()); vals = list(kpis["top_v"].values())
    ax.barh(vendors[::-1], vals[::-1], color="#2E7D32"); ax.set_xlabel("Purchase Amount"); ax.set_title("Top Vendors by Spend")
    fig.tight_layout(); vendor_path="chart_vendors.png"; fig.savefig(vendor_path, bbox_inches='tight', dpi=150); plt.close(fig); charts.append(vendor_path)
if kpis["top_m"]:
    fig, ax = plt.subplots(figsize=(8,4))
    mats = list(kpis["top_m"].keys()); qtys = list(kpis["top_m"].values())
    ax.bar(mats, qtys, color="#1565C0"); plt.xticks(rotation=45, ha='right'); ax.set_title("Top Materials by Quantity/Spend")
    fig.tight_layout(); mat_path="chart_materials.png"; fig.savefig(mat_path, bbox_inches='tight', dpi=150); plt.close(fig); charts.append(mat_path)
if kpis["monthly"]:
    fig, ax = plt.subplots(figsize=(9,3))
    months = list(kpis["monthly"].keys()); vals = list(kpis["monthly"].values())
    ax.plot(months, vals, marker='o'); ax.set_title("Monthly Purchase Trend"); ax.set_ylabel("Spend"); plt.xticks(rotation=45, ha='right')
    fig.tight_layout(); monthly_path="chart_monthly.png"; fig.savefig(monthly_path, bbox_inches='tight', dpi=150); plt.close(fig); charts.append(monthly_path)

# insert gauge first
charts.insert(0, gauge_path)

# UI cards (metric font tweak)
st.markdown("<style>.stMetric label{font-size:12px !important}</style>", unsafe_allow_html=True)
c1,c2,c3,c4,c5 = st.columns([1,1,1,1,1])
c1.metric("Total Records", kpis["records"])
c2.metric("Total Spend", f"{kpis['total_spend']:,.2f} {kpis.get('dominant','')}")
c3.metric("Dominant Currency", kpis.get('dominant','N/A'))
c4.metric("Procurement Risk", f"{risk['score']:.0f} ({risk['band']})")
c5.metric("Top Vendor", next(iter(kpis["top_v"]), "N/A"))

st.subheader("Visual Highlights")
for ch in charts:
    try:
        st.image(ch, use_container_width=True)
    except Exception:
        pass

# Show risk breakdown as a clean table, not a code blob
st.subheader("Procurement Risk – Breakdown")
bd_df = pd.DataFrame.from_dict(risk["breakdown"], orient="index", columns=["Score"]).reset_index().rename(columns={"index":"Metric"})
st.table(bd_df)

st.subheader("AI Insights")
ai_text = generate_ai(kpis)
# ensure AI summary is not truncated on the UI; show full
st.markdown(ai_text.replace("\n", "  \n"))

# Cover summary and PDF
summary_ai = generate_ai(kpis)[:1200]  # short snippet for cover
pdf_buffer = generate_pdf(ai_text, kpis, charts, company_name, summary_ai, risk)
st.download_button("📄 Download Full Executive Report (with Cover Page)", pdf_buffer, "SAP_Automatz_Executive_Report.pdf", "application/pdf")

# also offer risk CSV download for evidence
bd_csv = bd_df.to_csv(index=False).encode("utf-8")
st.download_button("📥 Download Risk Breakdown (CSV)", bd_csv, "risk_breakdown.csv", "text/csv")
